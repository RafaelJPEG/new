import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
import os

# Excel file setup
excel_file = 'students.xlsx'

if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.append(['Student Name', 'Course', 'Grade'])  # Header row
    wb.save(excel_file)

def submit_data():
    name = entry_name.get().strip()
    course = entry_course.get().strip()
    grade = entry_grade.get().strip()

    if not name or not course or not grade:
        messagebox.showerror("Error", "All fields are required.")
        return

    wb = load_workbook(excel_file)
    ws = wb.active
    ws.append([name, course, grade])
    wb.save(excel_file)

    entry_name.delete(0, tk.END)
    entry_course.delete(0, tk.END)
    entry_grade.delete(0, tk.END)

    messagebox.showinfo("Success", "Data submitted successfully!")

def view_data():
    wb = load_workbook(excel_file)
    ws = wb.active

    view_win = tk.Toplevel(root)
    view_win.title("View Data")

    tree = ttk.Treeview(view_win, columns=(1, 2, 3), show="headings")
    tree.heading(1, text="Student Name")
    tree.heading(2, text="Course")
    tree.heading(3, text="Grade")

    for row in ws.iter_rows(min_row=2, values_only=True):
        tree.insert('', tk.END, values=row)

    tree.pack(fill=tk.BOTH, expand=True)

# GUI Setup
root = tk.Tk()
root.title("Student Form")

tk.Label(root, text="Student Name:").grid(row=0, column=0, padx=10, pady=5, sticky='e')
tk.Label(root, text="Course:").grid(row=1, column=0, padx=10, pady=5, sticky='e')
tk.Label(root, text="Grade:").grid(row=2, column=0, padx=10, pady=5, sticky='e')

entry_name = tk.Entry(root)
entry_course = tk.Entry(root)
entry_grade = tk.Entry(root)

entry_name.grid(row=0, column=1, padx=10, pady=5)
entry_course.grid(row=1, column=1, padx=10, pady=5)
entry_grade.grid(row=2, column=1, padx=10, pady=5)

btn_submit = tk.Button(root, text="Submit", command=submit_data)
btn_view = tk.Button(root, text="View Data", command=view_data)

btn_submit.grid(row=3, column=0, columnspan=2, pady=10)
btn_view.grid(row=4, column=0, columnspan=2, pady=5)

root.mainloop()
