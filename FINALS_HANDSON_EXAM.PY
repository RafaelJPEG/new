import tkinter as tk
from tkinter import ttk
from openpyxl import Workbook, load_workbook

excel_file = 'students.xlsx'

try:
    wb = load_workbook(excel_file)
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    ws.append(['Name', 'Course', 'Grade'])
    wb.save(excel_file)

def submit_data():
    name = entry_name.get().strip()
    course = entry_course.get().strip()
    grade = entry_grade.get().strip()

    if not name or not course or not grade:
        print("Error: All fields are required.")
        return

    wb = load_workbook(excel_file)
    ws = wb.active
    ws.append([name, course, grade])
    wb.save(excel_file)

    entry_name.delete(0, tk.END)
    entry_course.delete(0, tk.END)
    entry_grade.delete(0, tk.END)

    print("Data submitted successfully!")

def view_data():
    wb = load_workbook(excel_file)
    ws = wb.active

    view_win = tk.Toplevel(root)
    view_win.title("View Data")

    tree = ttk.Treeview(view_win, columns=(1, 2, 3), show="headings")
    tree.heading(1, text="Name")
    tree.heading(2, text="Course")
    tree.heading(3, text="Grade")

    for row in ws.iter_rows(min_row=2, values_only=True):
        tree.insert('', tk.END, values=row)

    tree.pack(fill=tk.BOTH, expand=True)

root = tk.Tk()
root.title("Student Form")
root.configure(bg="lightgray")
root.geometry("400x350")

frame = tk.Frame(root, bg="lightgray")
frame.pack(expand=True, padx=20, pady=20)

title_label = tk.Label(frame, text="Student Form", font=("Arial", 16, "bold"), bg="lightgray", fg="green")
title_label.grid(row=0, column=0, columnspan=2, pady=20)

label_font = ("Arial", 12)
entry_font = ("Arial", 12)

tk.Label(frame, text="Name:", font=label_font, bg="lightgray").grid(row=1, column=0, padx=10, pady=5, sticky='e')
tk.Label(frame, text="Course:", font=label_font, bg="lightgray").grid(row=2, column=0, padx=10, pady=5, sticky='e')
tk.Label(frame, text="Grade:", font=label_font, bg="lightgray").grid(row=3, column=0, padx=10, pady=5, sticky='e')

entry_name = tk.Entry(frame, font=entry_font, width=25)
entry_course = tk.Entry(frame, font=entry_font, width=25)
entry_grade = tk.Entry(frame, font=entry_font, width=25)

entry_name.grid(row=1, column=1, padx=10, pady=5)
entry_course.grid(row=2, column=1, padx=10, pady=5)
entry_grade.grid(row=3, column=1, padx=10, pady=5)

btn_style = {"font": ("Arial", 12), "bg": "green", "fg": "white", "width": 15, "height": 2, "bd": 0, "relief": "flat"}
btn_submit = tk.Button(frame, text="Submit", command=submit_data, **btn_style)
btn_view = tk.Button(frame, text="View Data", command=view_data, **btn_style)

btn_submit.grid(row=4, column=0, columnspan=2, pady=15)
btn_view.grid(row=5, column=0, columnspan=2, pady=10)

root.mainloop()